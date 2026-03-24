import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from datetime import datetime, time, timedelta, date
from io import BytesIO

class ProductionGanttSystem:
    def __init__(self):
        self.START_HOUR = 3  # Day starts at 03:00 AM
        self.BASE_COL = 27   # Column AA
        self.COL_PER_DAY = 96
        self.DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
        
        # Mapping lines to specific rows
        self.LINE_ROWS = {
            "Pump": 11, "Ref-1": 14, "Ref-2": 17, 
            "Ref-3": 20, "Ref-4": 23, "Mini-1": 26
        }

    def get_monday_start(self, df):
        """Scans Column A for the first date and returns the Monday of that week."""
        for val in df.iloc[:, 0]:
            if isinstance(val, (datetime, date, pd.Timestamp)):
                # If it's a Timestamp/Datetime, convert to date
                d = val.date() if hasattr(val, 'date') else val
                # Calculate Monday (Weekday: Mon=0, Sun=6)
                return d - timedelta(days=d.weekday())
        return None

    def build_and_fill(self, fill_df):
        # 1. AUTO-DETECT START DATE
        start_date = self.get_monday_start(fill_df)
        if not start_date:
            raise ValueError("No valid date found in Column A of the Fill sheet.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Wk_{start_date.isocalendar()[1]}"
        
        # Styling Setup
        thin = Side(border_style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # 2. CONSTRUCT THE TEMPLATE
        for d_idx, day_name in enumerate(self.DAYS):
            col_start = self.BASE_COL + (d_idx * self.COL_PER_DAY)
            current_date = start_date + timedelta(days=d_idx)
            
            # Day Header
            ws.merge_cells(start_row=8, start_column=col_start, end_row=8, end_column=col_start + 95)
            day_cell = ws.cell(row=8, column=col_start, value=f"{day_name} {current_date}")
            day_cell.alignment = center_align
            day_cell.font = Font(bold=True, color="FFFFFF")
            day_cell.fill = PatternFill("solid", fgColor="4472C4") # Professional Blue
            
            # Hour Labels
            for h in range(24):
                hr_val = (self.START_HOUR + h) % 24
                c = col_start + (h * 4)
                ws.merge_cells(start_row=9, start_column=c, end_row=9, end_column=c + 3)
                hr_cell = ws.cell(row=9, column=c, value=f"{hr_val}:00")
                hr_cell.alignment = center_align
                hr_cell.border = border

        # 3. POPULATE DATA
        for idx, row in fill_df.iterrows():
            # Skip if row is empty or header-like
            if pd.isna(row[0]) or idx < 2: continue
            
            task_date = row[0]
            line_name = str(row[1]).strip()
            prod_name = str(row[2])
            start_t = row[4]
            end_t = row[5]

            if not isinstance(start_t, time) or not isinstance(end_t, time):
                continue

            # Calculate Day Offset
            actual_date = task_date.date() if hasattr(task_date, 'date') else task_date
            day_offset = (actual_date - start_date).days
            
            if day_offset < 0 or day_offset > 6: continue

            # Color Logic
            color = "D3D3D3" # Default Setup (Gray)
            if any(x in prod_name for x in ["DVB", "DV"]): color = "FFCC00" # Yellow
            elif "LSL" in prod_name: color = "99CCFF" # Blue
            
            sc = self.get_col_index(start_t, day_offset)
            ec = self.get_col_index(end_t, day_offset)
            target_row = self.LINE_ROWS.get(line_name, 11)
            
            fill = PatternFill("solid", fgColor=color)
            for col_idx in range(sc, ec):
                ws.cell(row=target_row, column=col_idx).fill = fill
                if col_idx == sc:
                    ws.cell(row=target_row, column=col_idx).value = prod_name
                    ws.cell(row=target_row, column=col_idx).font = Font(size=8)

        # Labels for the lines
        for name, r in self.LINE_ROWS.items():
            ws.cell(row=r, column=2, value=name).font = Font(bold=True)

        output = BytesIO()
        wb.save(output)
        return output.getvalue(), start_date

    def get_col_index(self, t, day_off):
        rel_h = t.hour - self.START_HOUR
        if t.hour < self.START_HOUR: rel_h += 24
        return self.BASE_COL + (day_off * self.COL_PER_DAY) + (rel_h * 4) + (t.minute // 15)

# --- Streamlit UI ---
st.set_page_config(page_title="Gantt Automator", layout="wide")
st.title("🏭 SGF Weekly Plan Creator")
st.info("Upload your **Fill.xlsm** file. The system will automatically detect the week and generate the Gantt chart.")

uploaded_file = st.file_uploader("Upload Fill_XX.xlsm", type=["xlsm"])

if uploaded_file:
    if st.button("🚀 Generate Weekly Schedule"):
        try:
            # Load only the 'Fill' sheet
            df = pd.read_excel(uploaded_file, sheet_name='Fill', header=None)
            
            engine = ProductionGanttSystem()
            excel_data, detected_date = engine.build_and_fill(df)
            
            st.success(f"✅ Week of **{detected_date}** detected and processed!")
            st.download_button(
                label="📥 Download Generated Excel",
                data=excel_data,
                file_name=f"Gantt_Schedule_{detected_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error: {e}")
